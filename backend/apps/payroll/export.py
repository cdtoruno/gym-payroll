import calendar
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril",
    "Mayo", "Junio", "Julio", "Agosto",
    "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

THIN  = Side(style="thin")
THICK = Side(style="medium")

LOGO_PATH = os.path.join(os.path.dirname(__file__), "gym_dado_logo.png")


def all_borders(thick=False):
    s = THICK if thick else THIN
    return Border(top=s, bottom=s, left=s, right=s)


HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
SUBHEAD_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
ALT_FILL     = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
TOTAL_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
FIRMA_FILL   = PatternFill("solid", start_color="F9F9F9", end_color="F9F9F9")

CURRENCY_FMT = '"C$"#,##0.00'
BODY_FONT    = "Calibri"
BODY_SIZE    = 11
HEADER_SIZE  = 11


def title_for_period(period, month, year):
    mes = MESES[month]
    if period == 1:
        label    = "1ra. Quincena"
        date_str = f"01/{month:02d}/{year} al 15/{month:02d}/{year}"
    else:
        last_day = calendar.monthrange(year, month)[1]
        label    = "2da. Quincena"
        date_str = f"16/{month:02d}/{year} al {last_day:02d}/{month:02d}/{year}"
    return f"Planilla {label} de {mes} del ({date_str})"


def build_sheet(ws, records, period, month, year):
    title = title_for_period(period, month, year)

    # ── Filas 1-3: cabecera con logo ──────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 24

    # Columna A filas 1-3: fondo azul para el logo
    ws.merge_cells("A1:A3")
    ws["A1"].fill = HEADER_FILL

    # Nombre empresa — columnas B a M, filas 1-2
    ws.merge_cells("B1:M2")
    ws["B1"] = "GYM DADO"
    ws["B1"].font      = Font(name=BODY_FONT, size=22, bold=True, color="FFFFFF")
    ws["B1"].fill      = HEADER_FILL
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    # Título del período — columnas B a M, fila 3
    ws.merge_cells("B3:M3")
    ws["B3"] = title
    ws["B3"].font      = Font(name=BODY_FONT, size=HEADER_SIZE, bold=True, color="FFFFFF")
    ws["B3"].fill      = SUBHEAD_FILL
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # Logo en esquina superior izquierda
    if os.path.exists(LOGO_PATH):
        img        = XLImage(LOGO_PATH)
        img.width  = 85
        img.height = 85
        img.anchor = "A1"
        ws.add_image(img)

    # ── Fila 4: espaciador ────────────────────────────────────────────────────
    ws.append([])
    ws.row_dimensions[4].height = 6

    # ── Fila 5: encabezados de grupo ──────────────────────────────────────────
    ws.merge_cells("F5:G5")
    ws["F5"] = "SALARIO"
    ws["F5"].font      = Font(name=BODY_FONT, size=HEADER_SIZE, bold=True, color="FFFFFF")
    ws["F5"].fill      = HEADER_FILL
    ws["F5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F5"].border    = all_borders(thick=True)

    ws.merge_cells("I5:K5")
    ws["I5"] = "Deducciones"
    ws["I5"].font      = Font(name=BODY_FONT, size=HEADER_SIZE, bold=True, color="FFFFFF")
    ws["I5"].fill      = HEADER_FILL
    ws["I5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["I5"].border    = all_borders(thick=True)
    ws.row_dimensions[5].height = 18

    # ── Fila 6: encabezados de columna ────────────────────────────────────────
    headers = [
        "No.",
        "Nombres y Apellidos",
        "Cargo",
        "Días\nLaborados",
        "Cédula",
        "Salario\nOrdinario",
        "Vacaciones",
        "Sub-Total\nDevengado",
        "Otras\nDeducciones",
        "Préstamo /\nAdelanto",
        "Desc.\nTardanzas",
        "Total\nDevengado",
        "Firma",
    ]
    ws.append(headers)
    for cell in ws[6]:
        cell.font      = Font(name=BODY_FONT, size=HEADER_SIZE, bold=True, color="FFFFFF")
        cell.fill      = SUBHEAD_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = all_borders()
    ws.row_dimensions[6].height = 32

    # ── Filas de datos ────────────────────────────────────────────────────────
    for idx, r in enumerate(records, start=1):
        fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
        row  = [
            idx,
            r["employee_name"],
            r["employee_position"],
            r["dias_laborados"],
            r["employee_cedula"],
            r["salary_base"],
            r["vacation_payment"],
            r["sub_total"],
            r["otras_deducciones"],
            r["prestamo_adelanto"],
            r.get("descuento_tardanzas", 0),
            r["total"],
            "",
        ]
        ws.append(row)
        data_row = ws[6 + idx]

        for ci, cell in enumerate(data_row, start=1):
            cell.font      = Font(name=BODY_FONT, size=BODY_SIZE)
            cell.fill      = fill
            cell.border    = all_borders()
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 4) else "left",
                vertical="center"
            )

        # Formato moneda
        for col_idx in (6, 7, 8, 9, 10, 11, 12):
            data_row[col_idx - 1].number_format = CURRENCY_FMT
            data_row[col_idx - 1].alignment     = Alignment(
                horizontal="right", vertical="center"
            )

        # Celda firma
        data_row[12].fill      = FIRMA_FILL
        data_row[12].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[6 + idx].height = 22

    # ── Fila de totales ───────────────────────────────────────────────────────
    total_row_num = 6 + len(records) + 1
    ws.append([])

    ws.merge_cells(
        start_row=total_row_num, start_column=1,
        end_row=total_row_num,   end_column=5
    )
    ws.cell(total_row_num, 1).value     = "TOTAL"
    ws.cell(total_row_num, 1).font      = Font(name=BODY_FONT, size=BODY_SIZE, bold=True)
    ws.cell(total_row_num, 1).fill      = TOTAL_FILL
    ws.cell(total_row_num, 1).border    = all_borders(thick=True)
    ws.cell(total_row_num, 1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    n = len(records)
    for col_idx, col_letter in [
        (6, "F"), (7, "G"), (8, "H"),
        (9, "I"), (10, "J"), (11, "K"), (12, "L")
    ]:
        cell = ws.cell(total_row_num, col_idx)
        cell.value         = f"=SUM({col_letter}7:{col_letter}{6 + n})"
        cell.number_format = CURRENCY_FMT
        cell.font          = Font(name=BODY_FONT, size=BODY_SIZE, bold=True)
        cell.fill          = TOTAL_FILL
        cell.border        = all_borders(thick=True)
        cell.alignment     = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[total_row_num].height = 22

    # ── Anchos de columna ─────────────────────────────────────────────────────
    widths = [12, 30, 16, 8, 22, 14, 14, 14, 14, 14, 12, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Configuración de impresión ────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = "1:6"


def generate_payroll_excel(records_data, period, month, year):
    """
    Genera el workbook completo.
    - Hoja General: todos los empleados
    - Una hoja por empleado
    """
    wb = Workbook()

    ws_general       = wb.active
    ws_general.title = "Todos los Empleados"
    build_sheet(ws_general, records_data, period, month, year)

    for r in records_data:
        parts      = r["employee_name"].split()
        sheet_name = " ".join(parts[:2])[:31]
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, [r], period, month, year)

    return wb